import os
import cv2
import pandas as pd
from fast_alpr import ALPR
from openpyxl import load_workbook
import tkinter as tk

# تحقق من وجود الفيديو
video_path = "video_2025-01-18_20-27-35.mp4"
if not os.path.exists(video_path):
    print(f"Error: Video file not found at {video_path}")
    exit()

# تحقق من وجود ملف Excel
excel_path = "car_numbers.xlsx"
if not os.path.exists(excel_path):
    print(f"Error: Excel file not found at {excel_path}")
    exit()

# قراءة ملف Excel
wb = load_workbook(excel_path)
ws = wb.active

# تهيئة ALPR
alpr = ALPR(
    detector_model="yolo-v9-t-384-license-plate-end2end",
    ocr_model="global-plates-mobile-vit-v2-model",
)

# فتح الفيديو
cap = cv2.VideoCapture(video_path)
if not cap.isOpened():
    print("Error: Could not open video.")
    exit()

# دالة للتحقق إذا كان رقم اللوحة موجود في ملف Excel
def check_plate_in_excel(plate_number):
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):  # نبدأ من الصف 2 لتجاهل العناوين
        if row[0].value == plate_number:
            return True
    return False

# دالة لتشغيل فيديو مختلف بناءً على الحالة
def play_video_once(video_path):
    cap = cv2.VideoCapture(video_path)
    while cap.isOpened():
        ret, frame = cap.read()
        if not ret:
            break
        cv2.imshow("Video", frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):  # إيقاف الفيديو عند الضغط على 'q'
            break
    cap.release()

# قراءة الإطارات ومعالجتها
try:
    video_played = False  # متغير للتحقق إذا تم تشغيل الفيديو بالفعل
    previous_plate_number = None  # لحفظ الرقم السابق الذي تم التحقق منه

    while True:
        ret, frame = cap.read()
        if not ret:
            # إعادة تشغيل الفيديو عند انتهائه
            cap.set(cv2.CAP_PROP_POS_FRAMES, 0)
            continue

        # الكشف عن اللوحات
        alpr_results = alpr.predict(frame)
        # استخراج الرقم وإضافته إلى متغير
        detected_plate_number = None

        for result in alpr_results:
            if hasattr(result, "ocr"):
                detected_plate_number = result.ocr.text
                break  # إيقاف التكرار إذا كنت تحتاج فقط إلى أول رقم

        # طباعة المتغير للتأكد
        print(f"Detected Plate Number: {detected_plate_number}")

        # تحقق من تطابق الرقم مع ملف Excel
        if detected_plate_number != previous_plate_number:  # تحقق إذا كان الرقم قد تغير
            previous_plate_number = detected_plate_number  # تحديث الرقم السابق
            video_played = False  # إعادة السماح بتشغيل الفيديو

        if detected_plate_number:
            if check_plate_in_excel(detected_plate_number) and not video_played:
                # إذا تطابق الرقم مع ملف Excel، تحديث اللون إلى الأخضر وتشغيل الفيديو الأخضر
                print("Plate number found in Excel, playing green video.")
                play_video_once("Untitled_Project_V1.mp4")
                video_played = True
            elif not check_plate_in_excel(detected_plate_number) and not video_played:
                # إذا لم يتطابق الرقم مع ملف Excel، تحديث اللون إلى الأحمر وتشغيل الفيديو الأحمر
                print("Plate number not found in Excel, playing red video.")
                play_video_once("Untitled_Project_V2 - Trim.mp4")
                video_played = True

        # رسم التنبؤات على الإطار
        annotated_frame = alpr.draw_predictions(frame)

        # عرض النتائج
        cv2.imshow("ALPR Video", annotated_frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):  # إيقاف الفيديو عند الضغط على 'q'
            break
except Exception as e:
    print(f"Error during processing: {e}")
finally:
    # تحرير الموارد
    cap.release()
    cv2.destroyAllWindows()

print("تم إيقاف تشغيل الفيديو.")
